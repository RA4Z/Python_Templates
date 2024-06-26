import json
import openpyxl

class Excel:
    def __init__(self, file: str):
        self.workbook = openpyxl.load_workbook(file, data_only=True)
        self.sheet = self.workbook['Database']

    def count_rows(self, column: int):
        max_row = self.sheet.max_row
        for row in range(max_row, 0, -1):
            cell_value = self.sheet.cell(row, column).value
            if cell_value is not None:
                return row
        return 0

    def get_cell(self, row: int, column: int):
        return self.sheet.cell(row, column).value

    def count_columns(self, row: int):
        max_col = self.sheet.max_column
        for col in range(max_col, 0, -1):
            cell_value = self.sheet.cell(row=row, column=col).value
            if cell_value is not None:
                return col
        return 0


if __name__ == "__main__":
    arquivo = ("\\\\intranet.weg.net@SSL\\DavWWWRoot\\br\\energia-wm\\pcp\\Comisso de Estoques\\Gráficos de giro de "
               "estoque\\CopyInventory.xlsm")
    excel = Excel(arquivo)
    rows = excel.count_rows(1)
    columns = excel.count_columns(1)
    header = [cell.value for cell in excel.sheet[1]]  # Primeira linha como cabeçalho
    data = []
    print(header)
    for index in range(2, rows):
        row_data = {}
        for col in range(1, columns + 1):
            row_data[header[col-1]] = excel.get_cell(index, col)
        data.append(row_data)

    with open("dados.json", "w", encoding="utf-8") as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=4)

    print("Arquivo JSON gerado com sucesso: dados.json")
